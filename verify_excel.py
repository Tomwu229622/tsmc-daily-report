from openpyxl import load_workbook
import os

EXCEL_PATH = r"C:\Users\K748\OneDrive - 財團法人中華民國對外貿易發展協會\FET\Stock分析\TSMC_股市分析報告.xlsx"
wb = load_workbook(EXCEL_PATH)
ws = wb["每日更新記錄"]
lr = ws.max_row
print(f"Max row: {lr}")
print(f"Last row A: {ws.cell(lr, 1).value}")
print(f"Last row B: {ws.cell(lr, 2).value}")
print(f"Last row C: {ws.cell(lr, 3).value}")
print(f"Last row F: {ws.cell(lr, 6).value}")

# Check HTML file
html_path = r"C:\Users\K748\OneDrive - 財團法人中華民國對外貿易發展協會\FET\Stock分析\TSMC_日報_2026-03-13.html"
exists = os.path.exists(html_path)
size = os.path.getsize(html_path) if exists else 0
print(f"HTML 2026-03-13 exists: {exists}, size: {size} bytes")
