"""
TSMC Excel 每日更新腳本 - 2026-07-01
執行：python update_excel_daily.py
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

EXCEL_PATH = r"C:\Users\K748\OneDrive - 財團法人中華民國對外貿易發展協會\FET\Stock分析\TSMC_股市分析報告.xlsx"

def make_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def style_cell(cell, bg="FFFFFF", align="center", bold=False, color="000000"):
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = make_border()

today_str = "2026-08-18"
tw_price = "NT$2,400（8/17收）"   # 台股2330 8/17 Mon官方收NT$2,400(+5、+0.21% vs 8/14 NT$2,395、開2,410/高2,420/低2,390、量13,482張TWSE口徑、成交額324.23億、成交56,197筆、TWSE確認)——⭐收復2,400整數關卡、止住8/14長黑，K線由長黑轉為小黑K帶上下影(實體僅10點、上影10、下影10)屬多空僵持型態；⚠️惟未收復5MA 2,408，更未挑戰2,415(8/12收)與2,425(8/13低=前報標記的假突破分水嶺)
change_pct = "+0.21%（8/17收）"    # 8/17官方+0.21%（TWSE；大盤同日+46.26點/+0.10%收45,857.27、盤中高46,189.30一度突破46,000後回落留長上影、成交額9,810.55億）——⚠️⚠️本日最關鍵訊號不在價格而在量能：成交量13,482張為2026年全年149個交易日最低量，第二低為8/11的18,248張、本日再低26.1%，僅及5日均量19,715張的68%；2,400的收復本質是「賣壓消失」而非「買盤回補」
nyse_price = "US$430.97（8/17收）" # NYSE TSM 8/17 Mon收$430.97(+$4.62、+1.08% vs 8/14 $426.35、Yahoo API＋stockanalysis.com雙來源確認、盤中15:55最後一根5分K為430.94)——隨費半反攻收紅(SOX+1.64%收12,621.00)、漲幅優於台股的+0.21%；⚠️惟距6/30歷史收盤高$477.57仍有-9.76%(2年日線查證、非創新高)；ADR溢價+14.21%(匯率31.80)
volume = "13,482（8/17收）"    # 台股8/17官方成交量13,482張（TWSE口徑）——⚠️⚠️為2026年全年149個交易日的最低量、僅及5日均量19,715張的68%、較8/14的21,163張萎縮36.3%；成交額324.23億、56,197筆
news_summary = '報告日2026-08-18(Tue)，涵蓋8/17(Mon)收盤。⚠️⭐量能乾涸下的止跌整理、籌碼罕見分歧：(1)⚠️⚠️本日最重要數據是量能——8/17成交量13,482張為2026年全年149個交易日最低量(第二低8/11的18,248張、本日再低26.1%)、僅及5日均量19,715張的68%，使2,400的收復缺乏成交量背書，屬「賣壓消失」而非「買盤回補」。(2)⭐籌碼面本波最極端分歧：外資自8/14的-3,024張反手大買+4,442張(本波單日最大買超)，證實8/14賣超屬MSCI被動調節的單日技術性行為；⚠️但投信同日賣超擴大至-2,058張(連2賣、近期最大)、自營+142張、三大合計+2,526張(T86官方)；全市場外資+454.47億、投信-177.77億、三大合計+290.76億(BFI82U官方)。(3)⭐⭐技術面關鍵確認：MACD DEA自-0.70上穿零軸至+1.20、DIF續揚至+8.76，快慢線雙雙站上零軸完成完整多頭排列(繼8/12 DIF上穿後的第二道確認)；⚠️惟紅柱連2日收縮(20.94→17.54→15.13)、KD出現死亡交叉(K 69.1下穿D 75.5、J自78.2降至56.1)——中期偏多、短期偏弱並存。(4)📊均線僅失守5MA 2,408(差8點)，10MA 2,388/20MA 2,364/60MA 2,374/120MA 2,196全數站穩，RSI 52.2續守中軸。(5)⭐美股半導體與大盤背離：SOX+1.64%收12,621.00但S&P500 -0.52%、那斯達克-0.32%、VIX+6.60%至15.19；設備股全面收復8/14失土(AMAT+5.55%收$535.31、LRCX+3.45%、ASML+2.12%、KLAC+1.00%)；⚠️客戶端落後(NVDA-0.07%、AVGO-0.14%、AMD-1.63%、QCOM-2.18%)。(6)⭐美光+4.13%收$1,011.75、逾6週來重返千元(前次為7/1的$1,032.28)；⚠️查證聲明：已用2年日線核對，8/17無任何相關個股創歷史新高(美光距6/25高$1,213.56有-16.63%、SOX距6/22高14,634.72有-13.76%)，市場流傳的「創新高」說法不予採用。(7)⚠️⚠️被低估的逆風：台幣自8/13的32.12四日升值1.00%至8/17的31.80(今晨31.78)，依法說指引每升1%毛利率約-40bps。(8)⚠️今晨夜盤倒V：TX 202608開46,000、衝高46,297後回落收45,811(-56、量25,712口)對現貨轉逆價差-46；CDF 202608夜盤收2,401(-2)對現貨僅+1。(9)⚠️明日8/19(週三)為8月台指期與台指選擇權「月」結算日，結算後未平倉與P/C比將重建、不可與結算前直接比較；8/17 P/C未平倉比117.16%(連3日回升)、CDF未平倉降至8,191口(連4日、屬轉倉)。(10)📊台股供應鏈極端分歧：欣興3037+9.80%收1,120、緯穎+2.87%、南亞科+2.34%；⚠️旺宏-5.47%、聯發科-3.80%收4,050、鴻海-1.73%。⚠️韓股8/17休市(光復節代替公休日)。今日觀察：量能能否回補至19,715張均量之上為第一順位、上檔5MA 2,408/2,415/2,425、下檔2,390/10MA 2,388/60MA 2,374。全程TWSE/TAIFEX/Yahoo官方API取數；技術指標依TWSE官方序列2026-01-02~08-17共149筆重算，並回算驗證8/14各值與前一日已發布日報完全一致。'

change_color = "00B050"        # green: 8/17 official +0.21% up

try:
    wb = load_workbook(EXCEL_PATH)
    ws = wb["每日更新記錄"]

    # 若最後一行日期 = 今日，更新該行；否則新增一行（避免重複 row）
    last_row = ws.max_row
    last_date = ws.cell(last_row, 1).value
    target_row = last_row if last_date == today_str else last_row + 1
    mode = "更新" if target_row == last_row else "新增"

    row_data = [
        today_str,
        tw_price,
        change_pct,
        nyse_price,
        volume,
        news_summary,
        "Yahoo Finance / Bloomberg"
    ]
    bg = "E9F2FB" if target_row % 2 == 0 else "FFFFFF"
    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=target_row, column=col, value=val)
        if col == 3:
            style_cell(c, bg=bg, align="center", color=change_color, bold=True)
        elif col == 6:
            style_cell(c, bg=bg, align="left")
        else:
            style_cell(c, bg=bg)

    ws["A2"] = f"最後更新：{today_str}"
    wb["封面總覽"]["A3"] = f"報告更新日期：{today_str}"
    wb.save(EXCEL_PATH)
    print(f"Excel {mode}成功：第 {target_row} 行（{today_str}）")
except Exception as e:
    print(f"Excel 更新失敗：{e}")
